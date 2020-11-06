# -*- coding: utf-8 -*-
"""
Created on Fri Nov  6 13:53:23 2020

@author: k-edit12a
"""

from openpyxl import load_workbook

wb = load_workbook('index.xlsx')

ws = wb.active

#a_sheet = wb.get_sheet_by_name('Sheet1')

index = ws['A2':'A3']

a = []

for row in index:
    for rows in row:
        a.append(rows.value)
        
import os
import os.path
import shutil

os.chdir('C:\Python_os\os2\os2')

A = ["109-2-全國學測-自然考科解析.pdf","109-2-全國學測-自然考科試題.pdf","109-2-全國學測-社會考科解析.pdf","109-2-全國學測-社會考科試題.pdf","109-2-全國學測-英文考科解析.pdf","109-2-全國學測-英文考科試題.pdf","109-2-全國學測-國文考科解析.pdf","109-2-全國學測-國文考科試題.pdf","109-2-全國學測-國文寫作考科解析.pdf","109-2-全國學測-國文寫作考科試題.pdf","109-2-全國學測-數學考科解析.pdf","109-2-全國學測-數學考科試題.pdf"]

B = []

for i in a:
    for j in A:
        B.append(i + '_' + '03' + '_' + j)

for k in A:
    for l in B:
        if k[11:19] == l[21:29]:
            shutil.copyfile(k,l)
            








